import openpyxl

def getData(path, sheetName):

    result=[]

    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetName]

    noOfRows=sheet.max_row
    noOfCols=sheet.max_column

    for r in range(2, noOfRows+1):
        temp=[]
        for c in range(1, noOfCols+1):
            temp.append(sheet.cell(r,c).value)
        result.append(temp)
    
    return result